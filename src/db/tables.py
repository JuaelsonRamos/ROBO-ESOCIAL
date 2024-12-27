from __future__ import annotations

from src.db import ClientConfig
from src.exc import Database
from src.gui.tkinter_global import TkinterGlobal
from src.sistema.sheet_constants import DEFAULT_MODEL_CELL, SHEET_FILETYPE_ASSOCIATIONS
from src.types import SheetModel

import re
import string
import hashlib

from datetime import datetime
from pathlib import Path
from typing import (
    Annotated,
    Any,
    ClassVar,
    Generic,
    Literal,
    Self,
    Sequence,
    TypeVar,
    TypedDict,
    cast,
    get_args,
)
from urllib.parse import urlparse

from openpyxl import (
    Workbook as _Workbook,
    load_workbook,
)
from openpyxl.cell.cell import Cell as _Cell
from openpyxl.worksheet.worksheet import Worksheet as _Worksheet
from sqlalchemy import (
    BLOB,
    INTEGER,
    REAL,
    TEXT,
    Boolean,
    DateTime,
    Enum,
    ForeignKey,
    MetaData,
    Row,
    TypeDecorator,
    delete,
    func,
    insert,
    select,
    sql,
    update,
)
from sqlalchemy.engine.interfaces import Dialect
from sqlalchemy.orm import DeclarativeBase, Mapped, MappedColumn, mapped_column


# region CUSTOM TYPES

RowDataType = int | float | str | bytes | bool | datetime


class timestamp(datetime): ...


class Url(TypeDecorator):
    impl = TEXT

    def process_literal_param(self, value: Any | None, dialect: Dialect) -> str:
        if value is None:
            return 'NULL'
        if not isinstance(value, str):
            raise TypeError(f'{type(value).__name__=} expected str or None')
        formatted = value.strip(string.whitespace)
        url = urlparse(formatted)
        if url.scheme == '' or url.netloc == '' or '.' not in url.netloc:
            raise ValueError(f'{url=} scheme or netloc invalid')
        return f"'{url.geturl()}'"  # literal string is single-quoted

    def process_result_value(self, value: Any | None, dialect: Dialect) -> Any | None:
        return urlparse(value) if value is not None else None


# endregion

# region DOCSTRING PARSING

re_whitespace = re.compile(f'[{string.whitespace}+]')


class SQLDocstring(TypedDict):
    doc: str
    comment: str


def doc(text: str) -> SQLDocstring:
    text = text.replace('\n', ' ')
    text = re_whitespace.sub(' ', text).strip(string.whitespace).strip('.').title()
    return {'doc': text, 'comment': text}


# endregion

# region BASE CLASS


class BaseDict(TypedDict, total=False):
    _id: int


TD = TypeVar('TD', bound=BaseDict)


class Base(DeclarativeBase, Generic[TD]):
    metadata = MetaData()
    type_annotation_map = {
        # standard sqlite3 type associaions
        int: INTEGER,
        bytes: BLOB,
        str: TEXT,
        float: REAL,
        # custom type associations
        datetime: DateTime,
        timestamp: REAL,
        Url: Url,
        bool: Boolean,
    }

    typed_dict: ClassVar[type[TD]]  # type: ignore

    # Common columns
    _id: Mapped[int] = mapped_column(
        autoincrement=True, primary_key=True, nullable=False, unique=True
    )

    @classmethod
    def sync_count(cls) -> int:
        with TkinterGlobal.sqlite.begin() as conn:
            query = func.count().select().select_from(cls)
            result = conn.execute(query).scalar_one_or_none()
            if result is None:
                return 0
            return result

    @classmethod
    def sync_select_all_ids(cls):
        with TkinterGlobal.sqlite.begin() as conn:
            query = select(cls._id)
            result = conn.scalars(query).all()
            return tuple(result)

    @classmethod
    def sync_delete_one_from_id(cls, _id: int):
        with TkinterGlobal.sqlite.begin() as conn:
            query = delete(cls).where(cls._id == _id)
            conn.execute(query)
            return _id

    @classmethod
    def sync_delete_many_from_id(cls, *row_ids: int):
        with TkinterGlobal.sqlite.begin() as conn:
            query = delete(cls).where(cls._id.in_(row_ids))
            conn.execute(query)
            return row_ids if isinstance(row_ids, tuple) else tuple(row_ids)

    @classmethod
    def sync_insert_one(cls, data: TD):
        with TkinterGlobal.sqlite.begin() as conn:
            query = insert(cls).values(**data)
            result = conn.execute(query)
            if result.inserted_primary_key is None:
                raise Database.InsertError
            return cast(int, result.inserted_primary_key[0])

    @classmethod
    def sync_insert_many(cls, *row_data: TD):
        with TkinterGlobal.sqlite.begin() as conn:
            query = insert(cls).values(*row_data)
            result = conn.execute(query)
            if len(result.inserted_primary_key_rows) == 0:
                raise Database.InsertError
            return tuple(cast(int, row[0]) for row in result)

    @classmethod
    def sync_select_one_from_id(cls, _id: int):
        with TkinterGlobal.sqlite.begin() as conn:
            query = select(cls).where(cls._id == _id)
            return conn.execute(query).one_or_none()

    @classmethod
    def sync_select_all_columns(cls, *columns: str):
        cols: list[MappedColumn[RowDataType]]
        try:
            cols = [getattr(cls, name) for name in columns]
        except AttributeError as err:
            raise Database.SelectError(err) from err
        with TkinterGlobal.sqlite.begin() as conn:
            query = select(*cols)
            result = conn.execute(query).all()
            tup: Annotated[tuple[Row[tuple[Self]], ...], len(columns)]
            tup = result if isinstance(result, tuple) else tuple(result)
            return tup

    @classmethod
    def sync_select_columns_from_id(
        cls, columns: Sequence[str], row_ids: Sequence[int]
    ):
        cols: list[MappedColumn[RowDataType]]
        try:
            cols = [getattr(cls, name) for name in columns]
        except AttributeError as err:
            raise Database.SelectError(err) from err
        with TkinterGlobal.sqlite.begin() as conn:
            query = select(*cols).where(cls._id.in_(row_ids))
            result = conn.execute(query).all()
            tup: Annotated[tuple[Row[tuple[Self]], ...], len(columns)]
            tup = result if isinstance(result, tuple) else tuple(result)
            return tup

    @classmethod
    def sync_select_many_from_id(cls, *row_ids: int):
        with TkinterGlobal.sqlite.begin() as conn:
            query = select(cls).where(cls._id.in_(row_ids))
            result = conn.execute(query).all()
            return result if isinstance(result, tuple) else tuple(result)

    @classmethod
    def sync_update_one_from_id(cls, data: TD, _id: int):
        with TkinterGlobal.sqlite.begin() as conn:
            query = update(cls).where(cls._id == _id).values(**data)
            conn.execute(query)
            return _id

    @classmethod
    def sync_update_many_from_id(cls, data: TD, *row_ids: int):
        with TkinterGlobal.sqlite.begin() as conn:
            query = update(cls).where(cls._id.in_(row_ids)).values(**data)
            conn.execute(query)
            return row_ids if isinstance(row_ids, tuple) else tuple(row_ids)


# endregion

BrowserType = Literal['firefox', 'chromium']
BrowserEnum = Enum(
    *get_args(BrowserType),
    name='browsertype',
    create_constraint=True,
    validate_strings=True,
)

SameSiteType = Literal['Strict', 'Lax', 'None']
SameSiteEnum = Enum(
    *get_args(SameSiteType),
    name='samesite',
    create_constraint=True,
    validate_strings=True,
)


class CommonColumns:
    @classmethod
    def created(cls):
        return mapped_column(
            nullable=False, unique=False, server_default=sql.func.now()
        )

    @classmethod
    def url(cls):
        return mapped_column(
            nullable=False,
            unique=False,
        )

    @classmethod
    def last_modified(cls):
        return mapped_column(
            nullable=True,
            unique=False,
            server_default=sql.null(),
            server_onupdate=sql.func.now(),
            **doc("""
                Datetime value is only non-null in case it has been modified at least
                once, while a creation date is always present
            """),
        )


class CookieDict(BaseDict, total=False):
    browser: BrowserType
    name: str
    value: str
    domain: str
    path: str
    expires: float
    http_only: bool
    secure: bool
    same_site: SameSiteType


class Cookie(Base):
    __tablename__ = 'cookie'
    typed_dict = CookieDict
    browser: Mapped[BrowserType] = mapped_column(
        BrowserEnum,
        nullable=False,
        unique=False,
        **doc('which playwright browser does these cookies are valid for'),
    )
    name: Mapped[str] = mapped_column(nullable=False, unique=False)
    value: Mapped[str] = mapped_column(nullable=False, unique=False, default='')
    domain: Mapped[Url] = mapped_column(nullable=False, unique=False)
    path: Mapped[str] = mapped_column(nullable=False, unique=False)
    expires: Mapped[float] = mapped_column(nullable=False, unique=False)
    http_only: Mapped[bool] = mapped_column(nullable=False, unique=False)
    secure: Mapped[bool] = mapped_column(nullable=False, unique=False)
    same_site: Mapped[SameSiteType] = mapped_column(
        SameSiteEnum, nullable=False, unique=False
    )


class LocalStorageDict(BaseDict, total=False):
    created: datetime
    last_modified: datetime | None
    origin_id: int
    name: str
    value: str


_local_storage_value_doc = """
    Length of name and value are theorically infinite as long as the total storage
    taken by name=value pairs doesn't exceed 5MiB, whether thats five 1MiB pairs
    (strings), or one 5MiB pair, doesn't matter
"""


class LocalStorage(Base):
    __tablename__ = 'localstorage'
    typed_dict = LocalStorageDict
    created: Mapped[datetime] = CommonColumns.created()
    last_modified: Mapped[datetime] = CommonColumns.last_modified()
    origin_id: Mapped[int] = mapped_column(
        ForeignKey('origin._id'), nullable=False, unique=False
    )
    name: Mapped[str] = mapped_column(
        nullable=False,
        unique=False,
        default='',
        **doc(_local_storage_value_doc),
    )
    value: Mapped[str] = mapped_column(
        nullable=False,
        unique=False,
        **doc(_local_storage_value_doc),
    )


class OriginDict(BaseDict, total=False):
    created: datetime
    origin: str


class Origin(Base):
    __tablename__ = 'origin'
    typed_dict = OriginDict
    created: Mapped[datetime] = CommonColumns.created()
    origin: Mapped[Url] = mapped_column(nullable=False, unique=False)


class Origin_LocalStorageDict(BaseDict, total=False):
    origin_id: int
    localstorage_id: int


_origin_localstorage_doc = doc('Association table: many origin to many local_storage')


class Origin_LocalStorage(Base):
    __tablename__ = 'origin_localstorage'
    typed_dict = Origin_LocalStorageDict
    __table_args__ = {'comment': _origin_localstorage_doc['comment']}
    __doc__ = _origin_localstorage_doc['doc']
    origin_id: Mapped[int] = mapped_column(
        ForeignKey('origin._id'), nullable=False, unique=False
    )
    localstorage_id: Mapped[int] = mapped_column(
        ForeignKey('localstorage._id'), nullable=False, unique=False
    )


class Cookie_BrowserContextDict(BaseDict, total=False):
    browsercontext_id: int
    cookie_id: int


_cookie_browsercontext_doc = doc(
    'Association table: many browsercontext to many cookie'
)


class Cookie_BrowserContext(Base):
    __tablename__ = 'cookie_browsercontext'
    typed_dict = Cookie_BrowserContextDict
    __table_args__ = {'comment': _cookie_browsercontext_doc['comment']}
    __doc__ = _cookie_browsercontext_doc['doc']
    browsercontext_id: Mapped[int] = mapped_column(
        ForeignKey('browsercontext._id'),
        nullable=False,
        unique=False,
    )
    cookie_id: Mapped[int] = mapped_column(
        ForeignKey('cookie._id'), nullable=False, unique=False
    )


class Origin_CookieDict(BaseDict, total=False):
    cookie_id: int
    origin_id: int


_origin_browsercontext_doc = doc('Association table: many cookie to many origin')


class Origin_Cookie(Base):
    __tablename__ = 'origin_cookie'
    typed_dict = Origin_CookieDict
    __table_args__ = {'comment': _origin_browsercontext_doc['comment']}
    __doc__ = _origin_browsercontext_doc['doc']
    cookie_id: Mapped[int] = mapped_column(
        ForeignKey('cookie._id'),
        nullable=False,
        unique=False,
    )
    origin_id: Mapped[int] = mapped_column(
        ForeignKey('origin._id'), nullable=False, unique=False
    )


CertificateOptions: tuple[str, ...] = ('PFX', 'CRT', 'PEM')
CertificateType = Literal[*CertificateOptions]
CertificateEnum = Enum(
    *get_args(CertificateType),
    name='certificate',
    create_constraint=True,
    validate_strings=True,
)


class ClientCertificateDict(BaseDict, total=False):
    created: datetime
    last_modified: datetime | None
    browsercontext_id: int | None
    origin: str
    cert_path: str | None
    cert: bytes | None
    key_path: bytes | None
    key: bytes | None
    pfx_path: str | None
    pfx: bytes | None
    passphrase: str | None
    using_type: str | None
    description: str


class ClientCertificate(Base):
    __tablename__ = 'clientcertificate'
    typed_dict = ClientCertificateDict
    created: Mapped[datetime] = CommonColumns.created()
    last_modified: Mapped[datetime] = CommonColumns.last_modified()
    browsercontext_id: Mapped[int] = mapped_column(
        ForeignKey('browsercontext._id'),
        nullable=True,
        unique=False,
        **doc('one client_certificate to many browsercontexts'),
    )
    origin: Mapped[str] = mapped_column(nullable=False, unique=False)
    cert_path: Mapped[str] = mapped_column(
        nullable=True, unique=True, server_default=sql.null()
    )
    cert: Mapped[bytes] = mapped_column(
        nullable=True, unique=True, server_default=sql.null()
    )
    key_path: Mapped[bytes] = mapped_column(
        nullable=True, unique=True, server_default=sql.null()
    )
    key: Mapped[bytes] = mapped_column(
        nullable=True, unique=True, server_default=sql.null()
    )
    pfx_path: Mapped[str] = mapped_column(
        nullable=True, unique=True, server_default=sql.null()
    )
    pfx: Mapped[bytes] = mapped_column(
        nullable=True, unique=True, server_default=sql.null()
    )
    passphrase: Mapped[str] = mapped_column(
        nullable=True, unique=False, server_default=sql.null()
    )
    using_type: Mapped[str] = mapped_column(
        CertificateEnum,
        nullable=True,
        unique=False,
        **doc('What kind of certificate is row refering to'),
    )
    description: Mapped[str] = mapped_column(
        nullable=False,
        unique=True,
        **doc('Unique user-defined description for that certificate'),
    )


ColorSchemeType = Literal['light', 'dark', 'no-preference', 'null']
ColorSchemeEnum = Enum(
    *get_args(ColorSchemeType),
    name='colorscheme',
    create_constraint=True,
    validate_strings=True,
)

ReducedMotionType = Literal['reduce', 'no-preference', 'null']
ReducedMotionEnum = Enum(
    *get_args(ReducedMotionType),
    name='reducedmotion',
    create_constraint=True,
    validate_strings=True,
)

ForcedColorsType = Literal['active', 'none', 'null']
ForcedColorsEnum = Enum(
    *get_args(ForcedColorsType),
    name='forcedcolors',
    create_constraint=True,
    validate_strings=True,
)

LocaleType = Literal['pt', 'pt-BR']
LocaleEnum = Enum(
    *get_args(LocaleType),
    name='locale',
    create_constraint=True,
    validate_strings=True,
)

TimezoneIdType = Literal[
    # Simple chart of timezone identifiers at: https://en.wikipedia.org/wiki/List_of_tz_database_time_zones
    # Some timezone identifiers link to others, e.g. 'America/Porto_Acre'
    # is linked to 'America/Rio_Branco'. Maybe because it's deprecated, since
    # it's listed in the 'backward' list, which I suppose stands for
    # 'Backwards Compatibility'? I didn't read through, just grabbed the values
    # because I know they are valid. I've written 'America/Recife' into a
    # couple of config files throughout time, and the timezone data has
    # always displayed correctly.
    'America/Araguaina',
    'America/Bahia',
    'America/Belem',
    'America/Boa_Vista',
    'America/Campo_Grande',
    'America/Cuiaba',
    'America/Eirunepe',
    'America/Fortaleza',
    'America/Maceio',
    'America/Manaus',
    'America/Noronha',
    'America/Porto_Velho',
    'America/Recife',
    'America/Rio_Branco',
    'America/Santarem',
    'America/Sao_Paulo',
]
TimezoneIdEnum = Enum(
    *get_args(TimezoneIdType),
    name='timezoneid',
    create_constraint=True,
    validate_strings=True,
)


class BrowserContextDict(BaseDict, total=False):
    created: datetime
    last_modified: datetime | None
    accept_downloads: bool
    offline: bool
    javascript_enabled: bool
    is_mobile: bool
    has_touch: bool
    colorscheme: ColorSchemeType
    reduced_motion: ReducedMotionType
    forced_colors: ForcedColorsType
    locale: LocaleType
    timezone_id: TimezoneIdType


class BrowserContext(Base):
    __tablename__ = 'browsercontext'
    typed_dict = BrowserContextDict
    created: Mapped[datetime] = CommonColumns.created()
    last_modified: Mapped[datetime] = CommonColumns.last_modified()
    accept_downloads: Mapped[bool] = mapped_column(nullable=False, unique=False)
    offline: Mapped[bool] = mapped_column(nullable=False, unique=False)
    javascript_enabled: Mapped[bool] = mapped_column(nullable=False, unique=False)
    is_mobile: Mapped[bool] = mapped_column(nullable=False, unique=False)
    has_touch: Mapped[bool] = mapped_column(nullable=False, unique=False)
    color_scheme: Mapped[ColorSchemeType] = mapped_column(
        ColorSchemeEnum, nullable=False, unique=False
    )
    reduced_motion: Mapped[ReducedMotionType] = mapped_column(
        ReducedMotionEnum, nullable=False, unique=False
    )
    forced_colors: Mapped[ForcedColorsType] = mapped_column(
        ForcedColorsEnum, nullable=False, unique=False
    )
    locale: Mapped[LocaleType] = mapped_column(LocaleEnum, nullable=False, unique=False)
    timezone_id: Mapped[TimezoneIdType] = mapped_column(
        TimezoneIdEnum, nullable=False, unique=False
    )


class ProcessingEntryDict(BaseDict, total=False):
    has_started: bool
    is_paused: bool
    has_finished: bool
    when_started: datetime | None
    when_last_started: datetime | None
    when_finished: datetime | None
    when_last_paused: datetime | None
    browsercontext_id: int
    workbook_id: int
    clientcertificate_id: int


class ProcessingEntry(Base):
    __tablename__ = 'processingentry'
    typed_dict = ProcessingEntryDict
    created: Mapped[datetime] = CommonColumns.created()
    last_modified: Mapped[datetime] = CommonColumns.last_modified()
    has_started: Mapped[bool] = mapped_column(
        nullable=False, unique=False, default=False
    )
    is_paused: Mapped[bool] = mapped_column(nullable=False, unique=False, default=False)
    has_finished: Mapped[bool] = mapped_column(
        nullable=False, unique=False, default=False
    )
    when_started: Mapped[datetime] = mapped_column(
        nullable=True, unique=False, server_default=sql.null()
    )
    when_last_started: Mapped[datetime] = mapped_column(
        nullable=True, unique=False, server_default=sql.null()
    )
    when_finished: Mapped[datetime] = mapped_column(
        nullable=True, unique=False, server_default=sql.null()
    )
    when_last_paused: Mapped[datetime] = mapped_column(
        nullable=True, unique=False, server_default=sql.null()
    )
    browsercontext_id: Mapped[int] = mapped_column(
        ForeignKey('browsercontext._id'), nullable=False, unique=False
    )
    workbook_id: Mapped[int] = mapped_column(
        ForeignKey('workbook._id'), nullable=False, unique=False
    )
    clientcertificate_id: Mapped[int] = mapped_column(
        ForeignKey('clientcertificate._id'), nullable=False, unique=False
    )


class EntryWorksheetDict(BaseDict, total=False):
    processingentry_id: int
    worksheet_id: int
    last_column: int
    last_row: int


class EntryWorksheet(Base):
    __tablename__ = 'entryworksheet'
    typed_dict = EntryWorksheetDict
    created: Mapped[datetime] = CommonColumns.created()
    last_modified: Mapped[datetime] = CommonColumns.last_modified()
    processingentry_id: Mapped[int] = mapped_column(
        ForeignKey('processingentry._id'), nullable=False, unique=False
    )
    worksheet_id: Mapped[int] = mapped_column(
        ForeignKey('worksheet._id'), nullable=False, unique=False
    )
    last_column: Mapped[int] = mapped_column(nullable=False, unique=False, default=0)
    last_row: Mapped[int] = mapped_column(nullable=False, unique=False, default=0)


ActionOfOriginType = Literal['processing', 'preview', 'recording']
ActionOfOriginEnum = Enum(
    *get_args(ActionOfOriginType),
    name='actionoforigin',
    create_constraint=True,
    validate_strings=True,
)


class ImageMediaDict(BaseDict, total=False):
    created: datetime
    blob: bytes
    sha512: str
    md5: str
    processingentry_id: int | None
    timeline_index: int | None
    action_of_origin: ActionOfOriginType
    dize: int
    width: int
    height: int


class ImageMedia(Base):
    __tablename__ = 'imagemedia'
    typed_dict = ImageMediaDict
    created: Mapped[datetime] = CommonColumns.created()
    blob: Mapped[bytes] = mapped_column(
        unique=False,
        nullable=False,
        **doc("""
            Blobs are part of a specific processing entry but two binary blobs may
            be equal by pure chance
        """),
    )
    sha512: Mapped[str] = mapped_column(unique=True, nullable=False)
    md5: Mapped[str] = mapped_column(unique=True, nullable=False)
    processingentry_id: Mapped[int] = mapped_column(
        ForeignKey('processingentry._id'),
        unique=False,
        nullable=True,
        **doc('One processingentry to many image_media'),
    )
    timeline_index: Mapped[int] = mapped_column(
        unique=False,
        nullable=True,
        **doc("""
            Index indicating it's position in the order of images created. Image 0 was
            the first created, and so on
        """),
    )
    action_of_origin: Mapped[ActionOfOriginType] = mapped_column(
        ActionOfOriginEnum,
        unique=False,
        nullable=False,
        **doc("""
            Action that originated the image, identical images may be originated from
            diferent actions by pure chance, but they wouldn't be treated differently,
            even if the difference could be noticed"
        """),
    )
    size: Mapped[int] = mapped_column(nullable=False, unique=False)
    width: Mapped[int] = mapped_column(nullable=False, unique=False)
    height: Mapped[int] = mapped_column(nullable=False, unique=False)


class WorksheetDict(BaseDict, total=False):
    created: datetime
    last_modified: datetime | None
    title: str
    workbook_index: int
    workbook_id: int
    dimensions: str
    columns: int
    rows: int
    mime_type: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    model_cell: str
    model_name: str
    model_code: int


class Worksheet(Base):
    __tablename__ = 'worksheet'
    typed_dict = WorksheetDict
    created: Mapped[datetime] = CommonColumns.created()
    last_modified: Mapped[datetime] = CommonColumns.last_modified()
    title: Mapped[str] = mapped_column(unique=False, nullable=False)
    workbook_index: Mapped[int] = mapped_column(unique=False, nullable=False)
    workbook_id: Mapped[int] = mapped_column(
        ForeignKey('workbook._id'), unique=False, nullable=False
    )
    dimensions: Mapped[str] = mapped_column(unique=False, nullable=False)
    columns: Mapped[int] = mapped_column(unique=False, nullable=False)
    rows: Mapped[int] = mapped_column(unique=False, nullable=False)
    mime_type: Mapped[str] = mapped_column(unique=False, nullable=False)
    min_row: Mapped[int] = mapped_column(unique=False, nullable=False)
    max_row: Mapped[int] = mapped_column(unique=False, nullable=False)
    min_col: Mapped[int] = mapped_column(unique=False, nullable=False)
    max_col: Mapped[int] = mapped_column(unique=False, nullable=False)
    model_cell: Mapped[str] = mapped_column(unique=False, nullable=False)
    model_name: Mapped[str] = mapped_column(unique=False, nullable=False)
    model_code: Mapped[int] = mapped_column(unique=False, nullable=False)

    @classmethod
    def from_sheet_obj(cls, book: _Workbook, sheet: _Worksheet) -> WorksheetDict:
        model_cell: _Cell = sheet[DEFAULT_MODEL_CELL]
        return WorksheetDict(
            title=sheet.title,
            workbook_index=book.index(sheet),
            dimensions=sheet.calculate_dimension(),
            columns=sheet.max_column,
            rows=sheet.max_row,
            mime_type=sheet.mime_type,
            min_row=sheet.min_row,
            max_row=sheet.max_row,
            min_col=sheet.min_column,
            max_col=sheet.max_column,
            model_cell=model_cell.coordinate,
            model_code=SheetModel.code_from_cell(model_cell),
            model_name=SheetModel.name_from_cell(model_cell),
        )


class WorkbookDict(BaseDict, total=False):
    created: datetime
    last_modified: datetime | None
    sha512: str
    md5: str
    blob: bytes
    epoch: datetime
    mime_type: str
    path: str
    template: bool
    excel_base_date: datetime | None
    file_type_suffix: str
    file_type_description: str
    file_size: int
    blob_size: int
    original_path: str


class Workbook(Base):
    __tablename__ = 'workbook'
    typed_dict = WorkbookDict
    created: Mapped[datetime] = CommonColumns.created()
    last_modified: Mapped[datetime] = CommonColumns.last_modified()
    sha512: Mapped[str] = mapped_column(unique=True, nullable=False)
    md5: Mapped[str] = mapped_column(unique=True, nullable=False)
    blob: Mapped[bytes] = mapped_column(unique=True, nullable=False)
    epoch: Mapped[datetime] = mapped_column(unique=False, nullable=False)
    mime_type: Mapped[str] = mapped_column(unique=False, nullable=False)
    path: Mapped[str] = mapped_column(unique=False, nullable=False)
    template: Mapped[bool] = mapped_column(unique=False, nullable=False)
    excel_base_date: Mapped[datetime] = mapped_column(unique=False, nullable=True)
    file_type_suffix: Mapped[str] = mapped_column(unique=False, nullable=False)
    file_type_description: Mapped[str] = mapped_column(unique=False, nullable=False)
    file_size: Mapped[int] = mapped_column(unique=False, nullable=False)
    blob_size: Mapped[int] = mapped_column(unique=False, nullable=False)
    original_path: Mapped[str] = mapped_column(unique=False, nullable=False)

    @classmethod
    def from_file(cls, path: Path) -> WorkbookDict:
        max_bytes = ClientConfig.SQLITE_LIMIT_LENGTH
        file_size = path.stat().st_size
        if file_size > max_bytes:
            raise Database.ValueError(f'file too big: {file_size=} > {max_bytes=}')
        blob = path.read_bytes()
        book = load_workbook(path, read_only=True)
        suffix = path.suffix.lower()
        desc: str | None = None
        for sheet_desc, suffix_globs in SHEET_FILETYPE_ASSOCIATIONS:
            if any(suffix == glob.lstrip('*').lower() for glob in suffix_globs):
                desc = sheet_desc
                break
        return WorkbookDict(
            sha512=hashlib.sha512(blob).hexdigest().upper(),
            md5=hashlib.md5(blob).hexdigest().upper(),
            blob=blob,
            epoch=book.epoch,
            mime_type=book.mime_type,
            path=book.path,
            template=book.template,
            excel_base_date=book.excel_base_date or None,
            file_type_suffix=suffix,
            file_type_description=desc or '',
            file_size=file_size,
            blob_size=len(blob),
            original_path=str(path),
        )
