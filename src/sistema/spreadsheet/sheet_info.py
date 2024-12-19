from __future__ import annotations

import string

from pathlib import Path
from typing import Sequence, TypeAlias

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


SheetFileSuffixOptions: TypeAlias = Sequence[tuple[str, tuple[str, ...]]]


def _make_file_dialog_options(pairs: SheetFileSuffixOptions) -> SheetFileSuffixOptions:
    """
    Utility function to allow list of options to grow, since is it defined in a standard
    format.

    Also, so that it (this function) can serve as a reference to implement this
    functinality elsewhere. The idea behind this function is arguable, yes, to it may be
    removed in the future, I genuenely think it may be useful for now, so let it be
    here, it's easy to substitute it, anyways.
    """
    options = []
    for desc, suffixes in pairs:
        formatted = tuple(s.lower() for s in suffixes)
        new_desc = '{} ({})'.format(desc, ';'.join(formatted))
        options.append((new_desc, suffixes))
    return tuple(options)


sheet_filetype_associations = (
    # NOTE: Excel 2010 can be read as .xls as well
    ('Excel 2010-presente', ('*.xlsx',)),
    # NOTE: openpyxl probably can't make anything out of macros
    ('Excel Macro 2010-presente', ('*.xlsm',)),
    # TODO: differ .xls file BIFF5, BIFF8, BIFF12
    # SEE: https://support.microsoft.com/en-us/office/file-formats-that-are-supported-in-excel-0943ff2c-6014-4e8d-aaea-b83d51d46247
    ('Excel 95-2003', ('*.xls',)),
    ('Excel XML 2003', ('*.xml',)),
    ('Planilha OpenDocument', ('*.ods',)),
    ('Arquivo CSV', ('*.csv',)),
    ('Arquivo JSON', ('*.json',)),
    ('Todos os arquivos', ('*',)),
)

sheet_filedialog_options = _make_file_dialog_options(sheet_filetype_associations)


class SheetInfo:
    worksheet: Worksheet
    columns: int
    rows: int
    dimensions: tuple[str, str]
    path: Path
    st_size: int
    st_created: float
    st_modified: float
    model_cell: Cell
    file_type_suffix: str
    file_type_description: str

    def __init__(self, path: str | Path) -> None:
        if isinstance(path, str):
            sheet = load_workbook(path).active
            path = Path(path)
        else:
            sheet = load_workbook(str(path)).active
        if sheet is None:
            raise ValueError(
                'path successfully parsed as workbook but it contains no worksheet'
            )
        self.path = path
        stat = self.path.stat()
        self.st_size = stat.st_size
        self.st_created = stat.st_ctime
        self.st_modified = stat.st_mtime
        self.worksheet = sheet
        self.model_cell = self.worksheet['A1']
        self.columns = self.worksheet.max_column
        self.rows = self.worksheet.max_row
        dimensions = self.worksheet.calculate_dimension()
        if dimensions == '' or ':' not in dimensions:
            self.dimensions = ('', '')
        else:
            self.dimensions = tuple(dimensions.split(':'))  # type: ignore
        self.file_type_description = ''
        self.file_type_suffix = ''
        for desc, suffixes in sheet_filetype_associations:
            normal_self_suffix = self.path.suffix.strip(string.punctuation).lower()
            normalized_suffixes = [
                s.strip(string.punctuation).lower() for s in suffixes
            ]
            if normal_self_suffix not in normalized_suffixes:
                continue
            self.file_type_description = desc.strip(string.whitespace)
            self.file_type_suffix = normal_self_suffix
            break
