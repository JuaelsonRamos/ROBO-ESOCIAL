from __future__ import annotations

from .exceptions import SpreadsheetObjectError
from .sheet_columns import SheetColumns
from .sheet_model import SheetModel

from sistema.models import Cell
from sistema.parsers import Modelo1, Parser

from pathlib import Path
from typing import Any, Never

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class Spreadsheet:
    def __init__(self, filename: str):
        assert isinstance(filename, str)
        path = Path(filename)
        assert path.exists()
        assert path.is_file()
        self.filetype = path.suffix.lstrip('.')
        assert self.filetype in ('xlsx', 'xls')
        self.path = path
        self.workbook = load_workbook(filename, read_only=True)
        assert isinstance(self.workbook.active, Worksheet)
        self.worksheet = self.workbook.active
        self.sheet_title = self.worksheet.title
        self.model = SheetModel(self.worksheet)
        self.columns = SheetColumns(self.worksheet)
        self.min_row: int = 3
        self.max_row: int = self.worksheet.max_row
        self.min_column: int = 1
        self.max_column: int = len(self.columns)

        self.parser: Parser
        match self.model.id:
            case 1:
                self.parser = Modelo1
            case 2 | _:
                raise NotImplementedError(
                    f"parser para '{self.model.name}' não implementado"
                )

    def cell_exists(self, row: int, column: int) -> bool:
        # Células inexistentes dentro desse intervalo são consideradas células vazias
        assert row > 0 and column > 0
        if row == 1 and column == 1:
            # célula do identificador do modelo
            return True
        elif row == 2 and column <= self.max_column:
            # célula de título de coluna
            return True
        elif row < self.min_row or column > self.max_column:
            return False
        return (
            self.min_row <= row <= self.max_row
            and self.min_column <= column <= self.max_column
        )

    def cell(self, row: int, column: int) -> Cell | Never:
        if not self.cell_exists(row, column):
            raise SpreadsheetObjectError('célula não existe')
        cell = self.worksheet.cell(row, column)
        return self.parser.parse_cell(self.columns[column], cell)

    def row(self, index: int) -> tuple[Cell, ...] | Never:
        if not (self.min_row <= index <= self.max_row):
            raise SpreadsheetObjectError('row não existe')
        row = next(self.worksheet.iter_rows(min_row=index, max_row=index))
        return self.parser.parse_row(self.columns, row)

    def walk(self):
        # TODO generator yields data pydantic model
        for row in self.worksheet.iter_rows(
            min_row=self.min_row,
            max_row=self.max_row,
            min_col=self.min_column,
            max_col=self.max_column,
        ):
            yield self.parser.parse_row(self.columns, row)
