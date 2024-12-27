from __future__ import annotations

import src.db.tables as table

from src.exc import SheetParsing
from src.sistema.sheet_data_schema import Modelo1Schema, Modelo2Schema
from src.sistema.validator import (
    CellModel,
    ColumnModel,
    Validator,
    cell_value_to_string,
)
from src.types import EmptyValueType, IsRequired

import gc
import io
import re
import string

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Generator, Never, Sequence, get_type_hints

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from typing_extensions import TypeIs
from unidecode import unidecode_expect_nonascii as unidecode


# TODO: generate column model
# TODO: validate cell


@dataclass(frozen=True, slots=True)
class WorkbookMetadata:
    _id: int
    created: datetime
    last_modified: datetime | None
    sha512: str
    md5: str
    epoch: datetime
    mime_type: str
    path: str
    template: bool
    excel_base_date: datetime | None
    file_type_suffix: str
    file_type_description: str
    file_size: int
    blob_size: int
    original_path: Path

    @classmethod
    def fetch_current(cls, db_id: int) -> WorkbookMetadata:
        fields = list(get_type_hints(table.WorkbookDict).keys())
        fields.append('_id')
        fields.remove('blob')
        rows = table.Workbook.sync_select_columns_from_id(fields, (db_id,))
        if len(rows) == 0:
            raise RuntimeError(f'workbook row {db_id=} does not exist in the database')
        data = rows[0]._asdict()
        data['original_path'] = Path(data['original_path'])
        return cls(**data)

    def fetch_file_blob(self) -> io.BytesIO:
        rows = table.Workbook.sync_select_columns_from_id(('blob',), (self._id,))
        if len(rows) == 0:
            raise RuntimeError(
                f'workbook row {self._id=} does not exist, although it should'
            )
        return io.BytesIO(rows[0].blob)


class WorksheetMetadata:
    _id: int
    created: datetime
    last_modified: datetime | None
    title: str
    workbook_index: int
    workbook_id: int
    dimensions: str
    columns: int
    rows: int
    mime_type: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    model_cell: str
    model_name: str
    model_code: int

    @classmethod
    def fetch_current(cls, db_id: int) -> WorksheetMetadata:
        data = table.Worksheet.sync_select_one_from_id(db_id)
        if data is None:
            raise RuntimeError(f'worksheet {db_id=} does not exist in the database')
        return cls(**data._asdict())


class SheetValidator:
    EmptyString: EmptyValueType = EmptyValueType()

    def __init__(self, workbook_id: int, worksheet_id: int) -> None:
        self.workbook_id = workbook_id
        self.workbook_metadata = WorkbookMetadata.fetch_current(workbook_id)
        self.worksheet_id = worksheet_id
        self.worksheet_metadata = WorksheetMetadata.fetch_current(worksheet_id)
        book_blob = self.workbook_metadata.fetch_file_blob()
        self.workbook = load_workbook(book_blob, read_only=True, rich_text=True)
        self.worksheet = self.workbook[self.worksheet_metadata.title]
        self.rows: int = self.worksheet_metadata.rows
        self.columns: int = self.worksheet_metadata.columns

        # force collect of workbook blob
        del book_blob
        gc.collect()

    def load_validation_schema(self) -> None:
        self.schema: Sequence[Validator]
        code = self.worksheet_metadata.model_code
        if code == 1:
            if not Modelo1Schema.is_loaded():
                Modelo1Schema.lazy_load_schema()
            self.schema = Modelo1Schema.get_schema()
            return
        if code == 2:
            if not Modelo2Schema.is_loaded():
                Modelo2Schema.lazy_load_schema()
            self.schema = Modelo2Schema.get_schema()
            return
        raise RuntimeError(f'unknown modelo code {code}')

    def as_bytes(self) -> io.BytesIO:
        with NamedTemporaryFile(delete=True) as tmp:
            self.workbook.save(tmp)
            return io.BytesIO(tmp.read())

    @classmethod
    def normalize_column_title(cls, value: str) -> str | EmptyValueType:
        value = value.strip(string.whitespace)
        if value == '':
            return cls.EmptyString
        value = unidecode(value)
        value = value.lower()
        replace = {'$': 's'}
        with io.StringIO() as property_name:
            for ch in value:
                property_name.write(replace.get(ch, ch))
            value = property_name.getvalue()
        value = re.sub(r'[^a-z0-9]+', '_', value).strip('_')
        if value == '':
            return cls.EmptyString
        return value

    @classmethod
    def is_empty_string(cls, value: Any) -> TypeIs[EmptyValueType]:
        return isinstance(value, EmptyValueType) and value is cls.EmptyString

    def cell_exists(self, row: int, column: int) -> bool:
        if row < 1 or column < 1:
            # indexes are positive and 1-based
            return False
        if (row, column) == (1, 1):
            # model metadata cell
            return True
        if row == 1 and column > 1:
            # there's only one valid cell in the first row: the very first cell
            return False
        headings: int = 2
        if row == headings and column <= self.columns:
            # column headings row
            # all cells are supposed to valid in this row
            return True
        # any other cell in the range have arbitrary values
        # if they have no value, they're deemed empty
        return headings <= row <= self.rows and 1 <= column <= self.columns

    def row_exists(self, index: int) -> bool:
        return 1 <= index <= self.rows

    def column_exists(self, index: int) -> bool:
        return 1 <= index <= self.columns

    def cell(self, row: int, column: int) -> Cell | Never:
        if not self.cell_exists(row, column):
            raise SheetParsing.IndexError(f'cell does not exist {row=} {column=}')
        return self.worksheet.cell(row, column)

    def row(self, index: int) -> tuple[Cell, ...] | Never:
        if not self.row_exists(index):
            raise SheetParsing.ValueError(f'row {index=} does not exist')
        row = next(self.worksheet.iter_rows(min_row=index, max_row=index))
        return row

    def column(self, index: int) -> tuple[Cell, ...] | Never:
        if not self.column_exists(index):
            raise SheetParsing.ValueError(f'column {index=} does not exist')
        col = next(self.worksheet.iter_cols(min_col=index, max_col=index))
        return col

    def headings(self):
        return self.row(2)

    def validate_and_load_columns(self) -> None:
        sequence: list[ColumnModel] = []
        for i, cell in enumerate(self.headings()):
            cell_validator = next(val for val in self.schema if val.matches(cell))
            text = cell_value_to_string(cell.value)
            obj = ColumnModel(
                index=i,
                original_text=text,
                required=IsRequired.from_cell(cell),
                validator=cell_validator,
            )
            sequence.append(obj)
        self.column_models: tuple[ColumnModel, ...] = tuple(sequence)

    def iter_row_cells(self) -> Generator[tuple[Cell, ...], None, None]:
        headings_index: int = 2
        return self.worksheet.iter_rows(
            min_row=headings_index + 1,
            max_row=self.rows,
            min_col=1,
            max_col=self.columns,
        )

    def iter_column_cells(self) -> Generator[tuple[Cell, ...], None, None]:
        headings_index: int = 2
        return self.worksheet.iter_cols(
            min_row=headings_index + 1,
            max_row=self.rows,
            min_col=1,
            max_col=self.columns,
        )

    def iter_and_validate_row_cell_models(
        self,
    ) -> Generator[tuple[CellModel], None, None | Never]:
        for row in self.iter_row_cells():
            if len(row) > len(self.column_models):
                raise SheetParsing.ValueError('more row cells than columns in sheet')
            model_sequence = []
            for i, cell in enumerate(row):
                column_model = self.column_models[i]
                validator_instance = column_model.validator.with_data(
                    column_model, cell
                )
                valid_cell_model = validator_instance.to_model()
                model_sequence.append(valid_cell_model)
            yield tuple(model_sequence)

    def iter_and_validate_column_cell_models(
        self,
    ) -> Generator[tuple[CellModel], None, None | Never]:
        for i, col in enumerate(self.iter_column_cells()):
            model_sequence = []
            for cell in col:
                column_model = self.column_models[i]
                validator_instance = column_model.validator.with_data(
                    column_model, cell
                )
                valid_cell_model = validator_instance.to_model()
                model_sequence.append(valid_cell_model)
            yield tuple(model_sequence)
