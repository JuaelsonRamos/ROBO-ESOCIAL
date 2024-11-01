from __future__ import annotations

from .fake_sheet import fake_sheet
from .iter import change_random

import pytest

import sistema.spreadsheet as sheet

import math

from openpyxl.workbook import Workbook


# Test
#   - args
#   - model name
#   - serialization

# DUMMY_TEXT = 'dummy_text_factory_meant'

# dummy_headers = [
#     Header(text='placeholder', required=Required.OPCIONAL, data_type='text'),
#     Header(text='placeholder', required=Required.OPCIONAL, data_type='float'),
#     Header(text='placeholder', required=Required.OPCIONAL, data_type='int'),
#     Header(text='placeholder', required=Required.OPCIONAL, data_type='date'),
#     Header(
#         text='placeholder',
#         required=Required.OPCIONAL,
#         data_type='factory',
#         factory=lambda: DUMMY_TEXT,
#     ),
# ]

dummy_headers = [
    {'text': 'placeholder_01', 'required': sheet.OPCIONAL},
    {'text': 'placeholder_02', 'required': sheet.OPCIONAL},
    {'text': 'placeholder_03', 'required': sheet.OPCIONAL},
    {'text': 'placeholder_04', 'required': sheet.OPCIONAL},
    {'text': 'placeholder_05', 'required': sheet.OPCIONAL},
    {'text': 'placeholder_06', 'required': sheet.OPCIONAL},
    {'text': 'placeholder_07', 'required': sheet.OPCIONAL},
    {'text': 'placeholder_08', 'required': sheet.OPCIONAL},
]


def test_args():
    # MODEL VERSION
    for i in range(-1, 10):
        if i in (1, 2):
            continue
        with pytest.raises(ValueError):
            with fake_sheet(model=i, headers=dummy_headers):
                pass

    with pytest.raises(ValueError):
        with fake_sheet(model=1, headers=[]):
            pass

    _nonsense: list[...] = [math.nan, 1, {}, 'string', float('inf'), range(99)]

    # HEADER LIST
    for error_value in _nonsense:
        with pytest.raises(ValueError):
            with fake_sheet(model=1, headers=change_random(dummy_headers, error_value)):
                pass

    # AMOUNT OF ROWS
    for i in range(-10, 0):
        with pytest.raises(ValueError):
            with fake_sheet(model=1, headers=dummy_headers, rows=i):
                pass

    # FAKER INSTANCE
    for error_value in _nonsense:
        with pytest.raises(ValueError):
            with fake_sheet(model=1, headers=dummy_headers, faker_instance=error_value):
                pass


def test_result():
    with fake_sheet(model=1, headers=dummy_headers, rows=1) as fs:
        assert isinstance(fs, Workbook)
        assert fs.active.max_column == len(dummy_headers)
        assert fs.active.max_row == 3


def test_cell_color():
    headers = [
        {'text': 'placeholder_01', 'required': sheet.OPCIONAL},
        {'text': 'placeholder_02', 'required': sheet.MAYBE},
        {'text': 'placeholder_03', 'required': sheet.REQUIRED},
        {'text': 'placeholder_04', 'required': sheet.MAYBE},
    ]

    expected = {
        'placeholder_01': sheet.Fill.WHITE,
        'placeholder_02': sheet.Fill.BLUE,
        'placeholder_03': sheet.Fill.RED,
        'placeholder_04': sheet.Fill.BLUE,
    }

    with fake_sheet(model=1, headers=headers) as fs:
        for i in range(1, len(headers)):
            cell = fs.active.cell(row=2, column=i)
            assert cell.value in expected
            assert expected[cell.value] == cell.fill
