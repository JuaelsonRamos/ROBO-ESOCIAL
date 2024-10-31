from __future__ import annotations

import sistema.spreadsheet as sheet

from contextlib import contextmanager
from typing import Iterable, Literal

from faker import Faker
from openpyxl import Workbook
from pydantic import TypeAdapter, ValidationError
from typing_extensions import TypedDict


class Header(TypedDict):
    text: str
    required: sheet.Requirement


@contextmanager
def fake_sheet(
    *,
    model: Literal[1, 2],
    headers: Iterable[Header],
    rows: int = 1,
    faker_instance: Faker | None = None,
):
    if model not in (1, 2):
        raise ValueError(f"modelo '{model}' é inválido")
    if rows < 1:
        raise ValueError('número de rows não pode ser menor que 1')
    if len(headers) == 0:
        raise ValueError('lista de headers está vazia')

    ta = TypeAdapter(Header)
    try:
        headers = tuple(ta.validate_python(h) for h in headers)
    except ValidationError:
        raise ValueError('elemento na lista de headers não é do tipo correto')

    if faker_instance is not None and not isinstance(faker_instance, Faker):
        raise ValueError('objeto não é uma instância de Faker')

    faker = faker_instance or Faker()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet['A1'] = f'Modelo {model}'

    for i, header in enumerate(headers, 1):
        cell = worksheet.cell(row=2, column=i, value=header['text'])
        cell.font = sheet.Font.HEADER
        cell.alignment = sheet.Alignment.HEADER

    for row in worksheet.iter_rows(
        min_row=3,  # delta
        max_row=rows + 2,  # different delta since row is always at least 1
        min_col=1,
        max_col=len(headers),
    ):
        for cell in row:
            cell.value = faker.sentence()
            cell.font = sheet.Font.CELL
            cell.alignment = sheet.Alignment.CELL

    try:
        yield workbook
    finally:
        workbook.close()
